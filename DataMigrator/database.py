from __future__ import annotations

import os
import re
import csv
from os import PathLike
from typing import Iterable, Callable, override

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from openpyxl.cell.cell import Cell
from openpyxl.cell.read_only import ReadOnlyCell


class Column:
    """ Column is a class that can load, store and operate data structure similar to column in
        mySQL (but without joints).

        # Attributes
        - title
        - comment (optional)
        - data (optional if it is a Iteration column such as index column)
        
        # Methods
        - count
        - add_data
        - get_data
        - swap_row
        - add_table
    """

    def __init__(self,
                 title: str,
                 comment: str | None = None,
                 data: list | Column | None = None,
                 mapping: dict | Callable = {"_Other": "_Origin"}
        ):
        """ Initialize a new Column with its title and other information.

            # Args
            - title
            - comment (optional)
            - data (optional): use given data to initialize the Column
            - mapping (optional): use a function of a mapping dict to adjust the initial data
            from the argment data. See About_rjson.md about how a mapping dict works.
        """
        self.title: str = title
        self.comment: str = comment if comment else ''
        self.data: list[object]
        
        if data:
            if isinstance(data, list):
                self.data = data.copy()
            elif isinstance(data, Column):
                self.data = data.data.copy()
            
            if isinstance(mapping, dict):
                self._mapping_dict(mapping)
            elif isinstance(mapping, Callable):
                self._mapping_func(mapping)
            
        else:
            self.data = []

    def __str__(self) -> str:
        return f"<Column \"{self.title}\" len(data) = {len(self.data)}>"
    
    def __repr__(self) -> str:
        return f"<Column \"{self.title}\" len(data) = {len(self.data)}>"
        
    def count(self) -> int:
        """ Get the length of data in this Column (number of rows). """
        return len(self.data)
    
    def add_data(self, newData) -> None:
        """ Add new data (row) at the bottom of the Column. """
        self.data.append(newData)

    def get_data(self, index: int) -> object | None:
        if index < len(self.data):
            return self.data[index]
        else:
            return None
        
    def del_row(self, index: int) -> None:
        """ Delete a row of given index. """
        if index < len(self.data):
            del self.data[index]
    
    def swap_row(self, i0: int, i1: int) -> None:
        """ Swap two rows by given the two indexes. """
        if i0 == i1:
            return
        if i0 > i1:
            i0, i1 = i1, i0
        if self.count() <= i1:
            self.data.extend([None] * (i1 - self.count() + 1))

        self.data[i0], self.data[i1] = self.data[i1], self.data[i0]
        
    def _mapping_func(self, func: Callable) -> None:
        """ Adjust all data in the column by passing them individually through the given
            function.
        """
        self.data = list(map(lambda s: None if s is None else func(s), self.data))
    
    def _mapping_dict(self, mapping: dict) -> None:
        """ Adjust all data in the column by a mapping dict. See About_rjson.md about how a
            mapping dict works.
        """
        for v in mapping.values():
            if v != "_Origin":
                break
        else:
            return
        
        for i in range(len(self.data)):
            if self.data[i] in mapping:
                m = mapping[self.data[i]]
            elif "_Other" in mapping:
                m = mapping["_Other"]

            if m != "_Origin":
                self.data[i] = m


class PlaceHolderColumn(Column):
    """ A placeholder Column for columns that suspend their process. """
    def __init__(self):
        Column.__init__(self, '')
    
    @override
    def swap_row(self, i0: int, i1: int) -> None:
        pass
    
    @override
    def del_row(self, index: int) -> None:
        pass

class EmptyColumn(Column):
    """ A empty Column with noting in it. """
    def __init__(self, title: str, comment: str | None = None):
        Column.__init__(self, title, comment)
    
    @override
    def count(self) -> int:
        return -1
    
    @override
    def get_data(self, index: int) -> None:
        return None
    
    @override
    def swap_row(self, i0: int, i1: int) -> None:
        pass
    
    @override
    def del_row(self, index: int) -> None:
        pass

class FilledColumn(Column):
    """ A filled Column with a constant value in all cells (rows) of it. """
    def __init__(self, title: str, comment: str | None = None, filler = None):
        Column.__init__(self, title, comment)
        self.filler = filler
    
    @override
    def count(self) -> int:
        return -1
    
    @override
    def get_data(self, index: int):
        return self.filler
    
    @override
    def swap_row(self, i0: int, i1: int) -> None:
        pass
    
    @override
    def del_row(self, index: int) -> None:
        pass
    
class IndexColumn(Column):
    """ A index Column marks the index of each row. """
    def __init__(self, title: str, comment: str | None = None, start_from: int = 1):
        Column.__init__(self, title, comment)
        self.start_from: int = start_from
    
    @override
    def count(self) -> int:
        return -1

    @override
    def get_data(self, index: int) -> int:
        if index < len(self.data):
            return self.data[index]
        return self.start_from + index
    
    @override
    def swap_row(self, i0: int, i1: int) -> None:
        for i in range(len(self.data)-1, max(i0, i1)):
            self.data.append(self.get_data[i])
        Column.swap_row(self, i0, i1)
    
    @override
    def del_row(self, index: int) -> None:
        for i in range(len(self.data), index+1):
            self.data.append(self.get_data(i))
        Column.del_row(self, index)
        self.start_from += 1


class Table:
    """ Table is a class that can load, store and operate data structure similar to table in
        mySQL (but without joints).

        # Attributes
        - name
        - columns: A list contains all columns in the Database.
        
        # Methods
        - import_from_xlsx (staticmethod)
        - export_to_xlsx
        - append_column
        - extend_columns
        - suspended_column
        - insert_column
        - index
        - get_range
        - get_column
        - del_row
        - swap_row
        - move_to_end
        - get_subtable
        - create_from_worksheet
    """

    def __init__(self, name: str, columns: list[Column] | None = None):
        """ Initialize a new table with its name. The columns argment is optional which allows
            you to initialize its data from a list of Column objects.
        """
        self.name: str = name
        self.columns: list[Column]
        self._column_index: dict[str, int]
        self._max_row_num: int

        if columns:
            self.columns = columns
            self._column_index = {c.title: i for i, c in enumerate(columns)}
            self._max_row_num = max([len(c) for c in columns])
        else:
            self.columns = []
            self._column_index = {}
            self._max_row_num = 0
    
    def __str__(self) -> str:
        return "<Table \"%s\" with %d columns and %d rows>" \
            % (self.name, len(self.columns), self._max_row_num)
    
    def __repr__(self) -> str:
        return "<Table \"%s\" with %d columns and %d rows>" \
            % (self.name, len(self.columns), self._max_row_num)
    
    def _check_new_title(self, new_title: str) -> None:
        """ Check if a title already exists in a table and raise a ValueError for an invalid
            title.
        """
        if new_title in self._column_index:
            raise ValueError("Title already exists.")
    
    def _check_empty_ends(self) -> bool:
        """ Check if there is any empty rows at the bottom of the Table, called by the method
            _clear_empty and not necessarily to call it manually.
        """
        c: Column
        for c in self.columns:
            if c.count() == 0:
                return False
            if len(c.data) < self._max_row_num:
                continue
            if c.data[-1] is not None:
                return False
        return True
    
    def _check_empty_row(self, rindex: int) -> bool:
        """ Check if a row is empty. """
        c: Column
        for c in self.columns:
            if len(c.data) < rindex:
                continue
            if c.data[rindex] is not None:
                return False
        return True
    
    def _pop_back(self) -> None:
        """ Clear and delete the last row of the Table, do not affect the columns that has
            less rows.
        """
        c: Column
        self._max_row_num -= 1
        for c in self.columns:
            if c.count() > self._max_row_num:
                del c.data[self._max_row_num:]
    
    def _clear_empty_ends(self) -> None:
        """ Clear empty rows at the end of the table, which is useful when imported from a
            worksheet with junks or empty cell at the bottom. Automatically called by the
            method create_from_worksheet and not necessarily to call it manually.
        """
        while self._check_empty_ends():
            self._pop_back()
    
    def _clear_empty_rows(self) -> None:
        for i in range(self._max_row_num-1, -1, -1):
            if self._check_empty_row(i):
                self.del_row(i)

    def _truncate_below_empty_row(self) -> None:
        """ If there is two empty rows in the table, truncate all rows below the empty row. """
        count = 0
        for i in range(self._max_row_num):
            if self._check_empty_row(i):
                count += 1
            else:
                count = 0
            if count >= 2:
                for i in range(i, self._max_row_num):
                    self._pop_back()
                break

    def append_column(self,
                      title: str,
                      comment: str | None,
                      column_type: type = Column,
                      **kwargs
        ) -> None:
        """ Append a new column at the end of the table. """
        new_column: Column = column_type(title, comment, **kwargs)
        
        self.columns.append(new_column)
        self._column_index[title] = len(self.columns) - 1
        self._max_row_num = max(new_column.count(), self._max_row_num)

    def extend_columns(self, content: Iterable[tuple[str, str]]) -> None:
        """ Add batch of columns at the end of the table, giving the titles and comments in
            the format of [(title, comment), (title, comment), ...]
        """
        for title, comment in content:
            self.append_column(title, comment)
    
    def suspended_column(self) -> int:
        """ Append a placeholder column at the end of the table and return the index of it """
        self.columns.append(PlaceHolderColumn())
        return len(self.columns) - 1
    
    def insert_column(self,
                      index: int,
                      title: str,
                      comment: str,
                      column_type: type = Column,
                      **kwargs
        ) -> None:
        """ Insert a new column in the specified place in a table. Empty Columns would be
            filled in the space.
        """
        if index >= len(self.columns):
            self.columns.extend([PlaceHolderColumn()] * (index - len(self.columns) + 1))
        
        new_column: Column = column_type(title, comment, **kwargs)
        self.columns[index] = new_column
        self._column_index[title] = index
        self._max_row_num = max(new_column.count(), self._max_row_num)
    
    def index(self, column_title: str) -> int:
        """ Returns the index of the column of the given title. Raise a KeyError if no such
            column in the table.
        """
        return self._column_index[column_title]
    
    def get_range(self, reverse=False) -> range:
        """ Returns an iterator that traversals the indexs of the table, can be from backwords
            if argment reverse == True.
        """
        if reverse:
            return range(self._max_row_num-1, -1, -1)
        return range(self._max_row_num)
    
    def get_column(self, column_title: str, loose: bool = False) -> Column:
        """ Returns the column of the given title. Raise a KeyError if no such column in the
            table.
        """
        if loose:
            for c in self.columns:
                if re.sub(r"\s", '', c.title) == re.sub(r"\s", '', column_title):
                    return c
                elif re.match(column_title, c.title):
                    return c
            raise KeyError(column_title)
        else:
            return self.columns[self._column_index[column_title]]
    
    def del_row(self, index: int) -> None:
        """ Delete a row of given index. """
        c: Column
        for c in self.columns:
            c.del_row(index)
        self._max_row_num -= 1
    
    def swap_row(self, i0: int, i1: int) -> None:
        """ Swap two rows of given index i0 and i1, affecting all columns in the table. """
        if i0 == i1:
            return
        if i0 > i1:
            i0, i1 = i1, i0
        if self._max_row_num <= i1:
            raise IndexError(f"row {i1} do not exists")
        
        c: Column
        for c in self.columns:
            c.swap_row(i0, i1)
    
    def move_to_end(self, index: int) -> None:
        """ Move the row to the end of the table, affecting all columns in the table. """
        if self._max_row_num <= index:
            raise IndexError(f"row {index} do not exists")
        if self._max_row_num - 1 == index:
            return
        
        for i in range(index, self._max_row_num - 1):
            self.swap_row(i, i + 1)
    
    def get_subtable(self, r0: int, c0: int, r1: int = -1, c1: int = -1) -> Table:
        """ Create and return a subtable from part of this table. """
        subtable: Table = Table("Subtable")
        if r1 == -1:
            r1 = self._max_row_num + 1
        if c1 == -1:
            c1 = len(self.columns)
        
        if r0 == 0:
            for c in range(c0, c1):
                title: str = self.columns[c].title
                data: list = self.columns[c].data[:r1-1]
                subtable.append_column(title, None, Column, data=data)
        else:
            for c in range(c0, c1):
                title: str = str(self.columns[c].data[r0-1])
                data: list = self.columns[c].data[r0:r1-1]
                subtable.append_column(title, None, Column, data=data)
        
        subtable._truncate_below_empty_row()
        return subtable
    
    @staticmethod
    def create_from_worksheet(ws: ReadOnlyWorksheet) -> Table:
        """ Create and return a Table object using the data from a worksheet object. """
        t: Table = Table(ws.title)
        
        r: tuple[ReadOnlyCell]
        for i, r in enumerate(ws.rows):

            if not t.columns: # first 
                c: ReadOnlyCell
                for j, c in enumerate(r):
                    t.columns.append(Column(c.value))
                    t._column_index[c.value] = j
                continue
            
            c: ReadOnlyCell
            for j, c in enumerate(r):
                t.columns[j].add_data(c.value)
        else:
            t._max_row_num = i
        
        t._clear_empty_ends()
        t._clear_empty_rows()
        return t
    
    @staticmethod
    def create_from_csv(rd: Iterable) -> Table:
        """ Create and return a Table object using the data from a csv reader object. """
        t: Table = Table("csv")
        
        r: list
        for i, r in enumerate(rd):

            if not t.columns: # first 
                for j, c in enumerate(r):
                    t.columns.append(Column(c))
                    t._column_index[c] = j
                continue
            
            for j, c in enumerate(r):
                t.columns[j].add_data(c)
        else:
            t._max_row_num = i
        
        t._clear_empty_ends()
        t._clear_empty_rows()
        return t


class Database:
    """ Database is a class that can load, store and operate data structure similar to mySQL
        (but without joints). Data can be imported from and exported to .xslx files (based on
        the openpyxl module).

        # Attributes
        - tables: A list contains all tables in the Database.
        
        # Methods
        - import_from_xlsx (staticmethod)
        - export_to_xlsx
        - has_table
        - get_table
        - add_table
    """

    def __init__(self):
        """ Initialize a Database with nothing in it. """
        self.tables: list[Table] = []
        self._table_index: dict[str, int] = {}

    @staticmethod
    def import_from_xlsx(path: PathLike) -> Database:
        """ Create and return a Database object using the data from a .xlsx file. """
        db: Database = Database()

        wb: Workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws: ReadOnlyWorksheet
        for i, ws in enumerate(wb.worksheets):
            new_table: Table = Table.create_from_worksheet(ws)
            db.tables.append(new_table)
            db._table_index[new_table.name] = i
        
        wb.close()
        
        return db

    @staticmethod
    def import_from_csv(path: PathLike) -> Database:
        """ Create and return a Database object using the data from a .csv file. """
        db: Database = Database()
        reader = csv.reader(open(path))
        db.tables.append(Table.create_from_csv(reader))
        return db
    
    @staticmethod
    def import_from_file(path: PathLike) -> Database:
        """ Create and return a Database object using the data from a supported file. """
        ext: str = os.path.splitext(path)[1]

        if ext in (".xls", ".xlsx"):
            return Database.import_from_xlsx(path)
        elif ext == ".csv":
            return Database.import_from_csv(path)
        else:
            raise Exception(f"Do not support {ext} file.")

    def export_to_xlsx(self, savePath: PathLike, add_comment: bool = False) -> None:
        """ Export the Database object to a .xlsx file. """
        wb: Workbook = Workbook(write_only=True)

        table: Table
        for table in self.tables:
            ws: WriteOnlyWorksheet = wb.create_sheet(table.name)
            ws.append([c.title for c in table.columns])
            if add_comment:
                ws.append([c.comment for c in table.columns])
            for i in range(table._max_row_num):
                ws.append([c.get_data(i) for c in table.columns])
        # print(savePath)
        wb.save(savePath)
    
    def has_table(self, name: str) -> bool:
        """ Return true if the Database has a table of the name. """
        return name in self._table_index
    
    def get_table(self, name: str, loose: bool = False) -> Table:
        """ Returns the table of the given name. Raise a KeyError if no such table in the
            database.
        """
        if loose:
            for t in self.tables:
                if re.sub(r"\s", '', t.name) == re.sub(r"\s", '', name):
                    return t
                elif re.match(name, t.name):
                    return t
            raise KeyError(name)
        else:
            return self.tables[self._table_index[name]]

    def add_table(self, name: str) -> Table:
        """ Add an empty table with nothing in it. """
        new_table: Table = Table(name)
        self.tables.append(new_table)
        self._table_index[name] = len(self.tables) - 1
        return new_table
